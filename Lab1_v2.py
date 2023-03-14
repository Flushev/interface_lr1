# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Lab1_v2.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.
import copy
import os
import shutil

from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QDoubleValidator, QIntValidator
from PyQt5.QtWidgets import QHeaderView, QItemDelegate, QStyledItemDelegate, QTableWidgetItem, QAbstractItemView, \
    QAbstractItemDelegate, QLineEdit
from qt_creator_packages.Lab1_v2_temp import Ui_MainWindow
from qt_creator_packages.Dialog import Ui_Dialog
from info.organizations import AVAILABLE_ORGANIZATIONS, META_DATA, AVAILABLE_ITEMS, ITEMS_DATA
from openpyxl import load_workbook


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, insex):
        return


class AllEditDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, insex):
        return


class Dialog(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()

        self.ui = Ui_Dialog()  # Экземпляр класса Ui_MainWindow, в нем конструктор всего GUI.
        self.ui.setupUi(self)  # Инициализация GUI

        self.ui.comboBox.setEditable(True)


class MyWin(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.ui = Ui_MainWindow()  # Экземпляр класса Ui_MainWindow, в нем конструктор всего GUI.
        self.ui.setupUi(self)  # Инициализация GUI

        self.dialog = Dialog()

        self.reworkTable()

        self.addComboToTable()

        self.setComboBoxEditable()

        self.setEvents()

        self.fieldCombo()

    def setEvents(self):
        # Открытие новое формв Dialog при нажатии ссылку "Ответственные лица"
        self.ui.pushButton_3.clicked.connect(self.openDialog)

        # Изменение списка подразделений в соответствие со справочником по выбранной организации
        self.ui.comboBox_2.currentTextChanged.connect(self.setComboBoxText)

        # Добавление строки
        self.ui.pushButton_2.clicked.connect(self.addRow)

        # Удаление строки
        self.ui.pushButton.clicked.connect(self.deleteRow)

        self.dialog.ui.comboBox.currentTextChanged.connect(self.setComboBoxTextDialog)

        self.ui.pushButton_4.clicked.connect(self.clearTable)

        self.ui.pushButton_5.clicked.connect(self.export)

    def openDialog(self):
        # pass
        self.dialog.ui.comboBox.clear()
        org_name = self.ui.comboBox_2.currentText()
        jobs = []
        for item in META_DATA:
            if item['organization_name'] == org_name:
                for person in item['response_persone']:
                    jobs.append(person['job'])
        self.dialog.ui.comboBox.addItems(jobs)

        self.dialog.exec_()

    def clearTable(self):
        while self.ui.tableWidget.rowCount() != 1:
            self.deleteRow()

    def export(self):
        if os.path.exists('result.xlsx'):
            os.remove('result.xlsx')
        shutil.copy2('form.xlsx', 'result.xlsx')

        fn1 = "result.xlsx"
        wb1 = load_workbook(fn1)
        ws1 = wb1["main"]

        # Организация
        cell_coord = 'A6'
        text = self.ui.comboBox_2.currentText()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Структурное подразделение
        cell_coord = 'A8'
        text = self.ui.comboBox.currentText()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # ОКПО
        cell_coord = 'BQ6'
        text = self.ui.lineEdit_11.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # ОКПД
        cell_coord = 'BQ9'
        text = self.ui.lineEdit_9.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Вид деятельности
        cell_coord = 'BQ10'
        text = self.ui.lineEdit_10.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Номер документа
        cell_coord = 'AP14'
        text = self.ui.lineEdit.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Дата составления
        cell_coord = 'AX14'
        text = self.ui.dateEdit.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Отчетный период с
        cell_coord = 'BF14'
        text = self.ui.dateEdit_2.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Отчетный период по
        cell_coord = 'BK14'
        text = self.ui.dateEdit_3.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Таблица
        if self.ui.tableWidget_2.rowCount() != 1:
            for i in range(self.ui.tableWidget.rowCount() - 1):
                if i <= 13:
                    delta = 23
                else:
                    delta = 32
                # Номер
                cell_coord = f'A{i + delta}'
                text = self.ui.tableWidget.item(i, 0).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Наименование
                cell_coord = f'E{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 1).currentText()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Код
                cell_coord = f'M{i + delta}'
                text = self.ui.tableWidget.item(i, 2).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Цена
                cell_coord = f'P{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 3).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Кол-во на начало отчетного периода
                cell_coord = f'U{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 4).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Сумма на начало отчетного периода
                cell_coord = f'Y{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 5).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Поступление за отчетный период их кладовой
                cell_coord = f'AD{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 6).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Сдано в кладовую и возмещено
                cell_coord = f'AK{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 7).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Кол-во на конец отчетного периода
                cell_coord = f'BI{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 8).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Сумма на конец отчетного периода
                cell_coord = f'BM{i + delta}'
                text = self.ui.tableWidget.cellWidget(i, 9).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

                # Примечание
                cell_coord = f'BS{i + delta}'
                text = self.ui.tableWidget.item(i, 10).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

            cells = ['U', 'Y', 'AD', 'AK', 'BI', 'BM']
            k = 0
            # Итого
            for i in [4, 5, 6, 7, 8, 9]:
                cell_coord = f'{cells[k]}{37}'
                text = self.ui.tableWidget.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                cell_coord = f'{cells[k]}{67}'
                text = self.ui.tableWidget.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                cell_coord = f'{cells[k]}{68}'
                text = self.ui.tableWidget.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                k += 1


            cells = ['AQ', 'AU', 'AZ', 'BD']
            k = 0
            # Итого
            for i in [0, 1, 2, 3]:
                cell_coord = f'{cells[k]}{37}'
                text = self.ui.tableWidget_2.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                cell_coord = f'{cells[k]}{67}'
                text = self.ui.tableWidget_2.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                cell_coord = f'{cells[k]}{68}'
                text = self.ui.tableWidget_2.item(self.ui.tableWidget.rowCount() - 1, i).text()
                self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)
                k += 1

        # Материальное лицо должность
        cell_coord = 'Q70'
        text = self.dialog.ui.comboBox.currentText()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Материальное лицо расшифровка
        cell_coord = 'AN70'
        text = self.dialog.ui.lineEdit.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        # Бухгалтер расшифровка
        cell_coord = 'V73'
        text = self.dialog.ui.lineEdit_2.text()
        self.insertCell(cell_coord=cell_coord, text=text, ws=ws1, wb=wb1, fn=fn1)

        wb1.close()

    def insertCell(self, cell_coord, text, ws, wb, fn):
        ws = ws
        wb = wb
        fn = fn
        cell_coord = cell_coord
        ws[cell_coord].value = text
        wb.save(fn)

    def setComboBoxTextDialog(self):
        name = ''
        org_name = self.ui.comboBox_2.currentText()
        for item in META_DATA:
            if item['organization_name'] == org_name:
                for person in item['response_persone']:
                    if person['job'] == self.dialog.ui.comboBox.currentText():
                        name = person['name']
        self.dialog.ui.lineEdit.clear()
        self.dialog.ui.lineEdit.setText(name)

    def setComboBoxText(self):
        # Привязка подразделений к организациям согласно справочнику
        self.ui.comboBox.clear()
        text = self.ui.comboBox_2.currentText()
        sub_organizations = []
        okpo = ''
        okpd = ''
        operation = ''
        for item in META_DATA:
            if item['organization_name'] == text:
                sub_organizations = copy.copy(item['sub_organizations'])
                okpo = item['okpo']
                okpd = item['okpd']
                operation = item['operation']
        self.ui.comboBox.addItems(sub_organizations)

        # Заполнение ОКПО, ОКДП и Вида деятельности
        self.ui.lineEdit_11.clear()
        self.ui.lineEdit_11.setText(okpo)
        self.ui.lineEdit_9.clear()
        self.ui.lineEdit_9.setText(okpd)
        self.ui.lineEdit_10.clear()
        self.ui.lineEdit_10.setText(operation)


    def fieldCombo(self):
        self.ui.comboBox_2.clear()
        self.ui.comboBox_2.addItems(AVAILABLE_ORGANIZATIONS)

    def reworkTable(self):
        # Таблица 1
        # Выравнивание первой вкладки таблицы
        for i in range(self.ui.tableWidget.columnCount()):
            self.ui.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)

        # Отображение вертикального заголовка (Итого)
        self.ui.tableWidget.verticalHeader().setVisible(True)
        self.ui.tableWidget.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

        # Таблица 2
        # Выравнивание второй вкладки таблицы
        for i in range(self.ui.tableWidget_2.columnCount()):
            self.ui.tableWidget_2.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
        self.ui.tableWidget_2.horizontalHeader().setFixedHeight(85)

        # Отображение вертикального заголовка (Итого)
        self.ui.tableWidget_2.verticalHeader().setVisible(True)



    def addComboToTable(self):
        # Создание выпадающего списка для наименований
        comboBox = QtWidgets.QComboBox()
        comboBox.setGeometry(QtCore.QRect(10, 50, 220, 22))
        comboBox.setSizeIncrement(QtCore.QSize(0, 4))
        comboBox.setObjectName("comboBox")
        comboBox.addItems(AVAILABLE_ITEMS)
        comboBox.setEditable(True)

        # Привязка события на выбор элемента к коду
        comboBox.currentTextChanged.connect(self.setTableComboText)
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 1, comboBox)

    def setTableComboText(self, text: str):
        item_code = ''
        for i in ITEMS_DATA:
            if i['item_name'] == text:
                item_code = i['item_code']
        item = QTableWidgetItem()
        item.setTextAlignment(Qt.AlignCenter)
        item.setText(item_code)
        self.ui.tableWidget.setItem(self.ui.tableWidget.rowCount() - 2, 2, item)

    def addRow(self):
        # Скрываем нумерацию строк в VerticalHeader и вставляем строку
        labels = []
        for i in range(self.ui.tableWidget.rowCount()):
            labels.append('')
        self.ui.tableWidget.insertRow(self.ui.tableWidget.rowCount() - 1)
        self.ui.tableWidget.setVerticalHeaderLabels(labels)
        self.ui.tableWidget_2.insertRow(self.ui.tableWidget_2.rowCount() - 1)
        self.ui.tableWidget_2.setVerticalHeaderLabels(labels)
        self.addComboToTable()



        # Делаем выравнивание по центру для элементов строки
        for i in range(self.ui.tableWidget.columnCount()):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            self.ui.tableWidget.setItem(self.ui.tableWidget.rowCount() - 2, i, item)
            # self.ui.tableWidget.itemChanged.connect(self.getTotal)

        for i in range(self.ui.tableWidget_2.columnCount()):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            self.ui.tableWidget_2.setItem(self.ui.tableWidget_2.rowCount() - 2, i, item)

        # Вставляем номер в первый столбец
        item = QTableWidgetItem()
        item.setTextAlignment(Qt.AlignCenter)
        item.setText(str(self.ui.tableWidget.rowCount() - 1))
        self.ui.tableWidget.setItem(self.ui.tableWidget.rowCount() - 2, 0, item)

        # Добавляем валидаторы на те столбцы, где это необходимо
        self.edit_double_list = []
        for i in range(5):
            edit_double = QLineEdit()
            edit_double.setValidator(QDoubleValidator(edit_double))
            edit_double.setAlignment(Qt.AlignHCenter)
            edit_double.setFrame(False)
            edit_double.textChanged.connect(self.getTotal)
            self.edit_double_list.append(edit_double)
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 3, self.edit_double_list[0])
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 5, self.edit_double_list[1])
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 9, self.edit_double_list[2])
        self.ui.tableWidget_2.setCellWidget(self.ui.tableWidget_2.rowCount() - 2, 1, self.edit_double_list[3])
        self.ui.tableWidget_2.setCellWidget(self.ui.tableWidget_2.rowCount() - 2, 3, self.edit_double_list[4])

        edit_int_list = []
        for i in range(6):
            edit_int = QLineEdit()
            edit_int.setValidator(QIntValidator(edit_int))
            edit_int.setAlignment(Qt.AlignHCenter)
            edit_int.setFrame(False)
            edit_int.textChanged.connect(self.getTotal)
            edit_int_list.append(edit_int)
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 4, edit_int_list[0])
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 6, edit_int_list[1])
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 7, edit_int_list[2])
        self.ui.tableWidget.setCellWidget(self.ui.tableWidget.rowCount() - 2, 8, edit_int_list[3])
        self.ui.tableWidget_2.setCellWidget(self.ui.tableWidget_2.rowCount() - 2, 0, edit_int_list[4])
        self.ui.tableWidget_2.setCellWidget(self.ui.tableWidget_2.rowCount() - 2, 2, edit_int_list[5])

        combo = self.ui.tableWidget.cellWidget(self.ui.tableWidget.rowCount() - 2, 1)
        self.setTableComboText(text=combo.currentText())

    def deleteRow(self):
        self.ui.tableWidget.removeRow(self.ui.tableWidget.rowCount() - 2)

        self.ui.tableWidget_2.removeRow(self.ui.tableWidget_2.rowCount() - 2)

        self.getTotal()

    def getTotal(self):
        # Расчет итогов на первой таблице
        for j in [4, 5, 6, 7, 8, 9]:
            total = 0
            for i in range(self.ui.tableWidget.rowCount() - 1):
                edit = self.ui.tableWidget.cellWidget(i, j)
                text = str(edit.text())
                if text is None or text == '' or text == ' ':
                    text = 0
                else:
                    text = text.replace(',', '.')

                total += float(text)
            self.ui.tableWidget.item(self.ui.tableWidget.rowCount() - 1, j).setText(str(total))

        # Расчет итогов на второй таблице
        for j in [0, 1, 2, 3]:
            total = 0
            for i in range(self.ui.tableWidget_2.rowCount() - 1):
                edit = self.ui.tableWidget_2.cellWidget(i, j)
                text = str(edit.text())
                if text is None or text == '' or text == ' ' or text == ',':
                    text = 0
                else:
                    text = text.replace(',', '.')

                total += float(text)
            self.ui.tableWidget_2.item(self.ui.tableWidget_2.rowCount() - 1, j).setText(str(total))

    def setComboBoxEditable(self):
        self.ui.comboBox.setEditable(True)
        self.ui.comboBox_2.setEditable(True)




